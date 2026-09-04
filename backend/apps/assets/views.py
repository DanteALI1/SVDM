import csv
import io

from django.contrib.auth import get_user_model
from django_filters import rest_framework as filters
from openpyxl import load_workbook
from rest_framework import serializers, viewsets, status, parsers, views
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from apps.tenants.permissions import IsTenantMember, IsTenantAdmin, IsAnalystOrAbove
from .models import Asset, AssetType, BusinessSystem, DEFAULT_ASSET_TYPES

User = get_user_model()


class AssetTypeSerializer(serializers.ModelSerializer):
    class Meta:
        model = AssetType
        fields = ["id", "code", "name", "is_system"]


class BusinessSystemSerializer(serializers.ModelSerializer):
    class Meta:
        model = BusinessSystem
        fields = ["id", "name", "code", "description"]


class AssetSerializer(serializers.ModelSerializer):
    asset_type_name = serializers.CharField(source="asset_type.name", read_only=True)
    business_system_name = serializers.CharField(source="business_system.name", read_only=True)
    owner_username = serializers.CharField(source="owner.username", read_only=True)
    contour_name = serializers.CharField(source="contour.name", read_only=True)

    class Meta:
        model = Asset
        fields = [
            "id",
            "name",
            "asset_type",
            "asset_type_name",
            "status",
            "ip_address",
            "fqdn",
            "os_platform",
            "environment",
            "criticality",
            "owner",
            "owner_username",
            "business_system",
            "business_system_name",
            "description",
            "tags",
            "inventory_number",
            "location",
            "contour",
            "contour_name",
            "security_officer",
            "commissioned_at",
            "created_at",
            "updated_at",
        ]


class AssetFilter(filters.FilterSet):
    environment = filters.CharFilter()
    status = filters.CharFilter()
    criticality = filters.CharFilter()
    asset_type = filters.NumberFilter()
    business_system = filters.NumberFilter()
    owner = filters.NumberFilter()
    group_by = filters.CharFilter(method="noop")

    class Meta:
        model = Asset
        fields = ["environment", "status", "criticality", "asset_type", "business_system", "owner"]

    def noop(self, qs, name, value):
        return qs


class AssetViewSet(viewsets.ModelViewSet):
    serializer_class = AssetSerializer
    permission_classes = [IsAuthenticated, IsTenantMember]
    filterset_class = AssetFilter
    search_fields = ["name", "fqdn", "ip_address", "inventory_number"]
    ordering_fields = ["name", "criticality", "environment", "commissioned_at"]

    def get_queryset(self):
        return Asset.objects.filter(tenant=self.request.tenant).select_related(
            "asset_type", "business_system", "owner", "contour", "security_officer"
        )

    def perform_create(self, serializer):
        serializer.save(tenant=self.request.tenant)

    def get_permissions(self):
        if self.action in ("create", "update", "partial_update", "destroy", "import_csv", "import_excel"):
            return [IsAuthenticated(), IsTenantMember(), IsAnalystOrAbove()]
        return super().get_permissions()

    @action(detail=False, methods=["get"])
    def grouped(self, request):
        group_by = request.query_params.get("by", "business_system")
        qs = self.filter_queryset(self.get_queryset())
        groups = {}
        for asset in qs:
            if group_by == "type":
                key = asset.asset_type.name
            elif group_by == "owner":
                key = asset.owner.username
            elif group_by == "environment":
                key = asset.environment
            else:
                key = asset.business_system.name
            groups.setdefault(key, []).append(AssetSerializer(asset).data)
        # Hierarchy: Org → System → Asset
        return Response({"tenant": request.tenant.name, "groups": groups})

    @action(detail=False, methods=["post"], parser_classes=[parsers.MultiPartParser])
    def import_csv(self, request):
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "file required"}, status=400)
        decoded = f.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        created = self._import_rows(reader, request)
        return Response({"created": created})

    @action(detail=False, methods=["post"], parser_classes=[parsers.MultiPartParser])
    def import_excel(self, request):
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "file required"}, status=400)
        wb = load_workbook(f, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: row[i] for i in range(len(headers)) if headers[i]})
        created = self._import_rows(rows, request)
        return Response({"created": created})

    def _import_rows(self, rows, request):
        created = 0
        tenant = request.tenant
        for row in rows:
            atype, _ = AssetType.objects.get_or_create(
                tenant=tenant,
                code=str(row.get("asset_type") or row.get("type") or "server").lower(),
                defaults={"name": str(row.get("asset_type") or "Server")},
            )
            bsys, _ = BusinessSystem.objects.get_or_create(
                tenant=tenant, name=str(row.get("business_system") or "Default")
            )
            from apps.tenants.models import Contour

            contour, _ = Contour.objects.get_or_create(
                tenant=tenant,
                code=str(row.get("contour") or "intranet").lower(),
                defaults={"name": str(row.get("contour") or "Intranet")},
            )
            owner = User.objects.filter(username=row.get("owner") or request.user.username).first() or request.user
            officer = (
                User.objects.filter(username=row.get("security_officer") or request.user.username).first()
                or request.user
            )
            Asset.objects.create(
                tenant=tenant,
                name=str(row["name"]),
                asset_type=atype,
                status=str(row.get("status") or "in_service"),
                ip_address=str(row["ip_address"]),
                fqdn=str(row["fqdn"]),
                os_platform=str(row.get("os_platform") or "unknown"),
                environment=str(row.get("environment") or "prod"),
                criticality=str(row.get("criticality") or "Medium"),
                owner=owner,
                business_system=bsys,
                description=str(row.get("description") or ""),
                tags=str(row.get("tags") or "").split(",") if row.get("tags") else [],
                inventory_number=str(row.get("inventory_number") or ""),
                location=str(row.get("location") or ""),
                contour=contour,
                security_officer=officer,
                commissioned_at=row.get("commissioned_at") or timezone_today(),
            )
            created += 1
        return created


def timezone_today():
    from django.utils import timezone

    return timezone.now().date()


class AssetTypeViewSet(viewsets.ModelViewSet):
    serializer_class = AssetTypeSerializer
    permission_classes = [IsAuthenticated, IsTenantMember]

    def get_queryset(self):
        return AssetType.objects.filter(tenant=self.request.tenant)

    def perform_create(self, serializer):
        if self.request.membership.role != "admin":
            from rest_framework.exceptions import PermissionDenied

            raise PermissionDenied()
        serializer.save(tenant=self.request.tenant, is_system=False)


class BusinessSystemViewSet(viewsets.ModelViewSet):
    serializer_class = BusinessSystemSerializer
    permission_classes = [IsAuthenticated, IsTenantMember]

    def get_queryset(self):
        return BusinessSystem.objects.filter(tenant=self.request.tenant)

    def perform_create(self, serializer):
        serializer.save(tenant=self.request.tenant)
