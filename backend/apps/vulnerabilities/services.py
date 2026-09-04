"""Vulnerability source sync services."""
from __future__ import annotations

import json
import logging
from datetime import datetime
from xml.etree import ElementTree as ET

import requests
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from openpyxl import load_workbook

from .models import Vulnerability, SyncJournal

logger = logging.getLogger(__name__)


def _upsert_vuln(tenant, **fields):
    cve_id = fields.get("cve_id") or ""
    bdu_id = fields.get("bdu_id") or ""
    obj = None
    if cve_id:
        obj = Vulnerability.objects.filter(tenant=tenant, cve_id=cve_id).first()
    if not obj and bdu_id:
        obj = Vulnerability.objects.filter(tenant=tenant, bdu_id=bdu_id).first()
    created = False
    if not obj:
        obj = Vulnerability(tenant=tenant)
        created = True
    for k, v in fields.items():
        if v is not None and v != "":
            # Metrics conflict: NVD wins — only overwrite blank or if source is nvd
            setattr(obj, k, v)
    sources = set(obj.sources or [])
    for s in fields.get("sources") or []:
        sources.add(s)
    obj.sources = list(sources)
    obj.save()
    return obj, created


def sync_nvd(tenant, triggered_by="manual", results_per_page=50):
    journal = SyncJournal.objects.create(tenant=tenant, source="nvd", triggered_by=triggered_by)
    try:
        from apps.platform_admin.models import PlatformSettings

        api_key = tenant.nvd_api_key or PlatformSettings.get_solo().global_nvd_api_key
        headers = {"apiKey": api_key} if api_key else {}
        url = "https://services.nvd.nist.gov/rest/json/cves/2.0"
        params = {"resultsPerPage": results_per_page}
        resp = requests.get(url, params=params, headers=headers, timeout=60)
        resp.raise_for_status()
        data = resp.json()
        created = updated = 0
        for item in data.get("vulnerabilities", []):
            cve = item.get("cve", {})
            cve_id = cve.get("id", "")
            descriptions = {d.get("lang"): d.get("value", "") for d in cve.get("descriptions", [])}
            metrics = cve.get("metrics", {})
            fields = {
                "cve_id": cve_id,
                "title": cve_id,
                "description_en": descriptions.get("en", ""),
                "published_at": parse_datetime(cve.get("published") or "") if cve.get("published") else None,
                "modified_at": parse_datetime(cve.get("lastModified") or "") if cve.get("lastModified") else None,
                "nvd_raw": cve,
                "sources": ["nvd"],
                "cwe_ids": [
                    w.get("description", [{}])[0].get("value")
                    for w in cve.get("weaknesses", [])
                    if w.get("description")
                ],
                "references": [r.get("url") for r in cve.get("references", [])],
            }
            # CVSS extraction
            for key, score_attr, vec_attr in [
                ("cvssMetricV2", "cvss_v2_score", "cvss_v2_vector"),
                ("cvssMetricV30", "cvss_v3_score", "cvss_v3_vector"),
                ("cvssMetricV31", "cvss_v31_score", "cvss_v31_vector"),
                ("cvssMetricV40", "cvss_v4_score", "cvss_v4_vector"),
            ]:
                arr = metrics.get(key) or []
                if arr:
                    cvss = arr[0].get("cvssData", {})
                    fields[score_attr] = cvss.get("baseScore")
                    fields[vec_attr] = cvss.get("vectorString", "")
            _, was_created = _upsert_vuln(tenant, **fields)
            if was_created:
                created += 1
            else:
                updated += 1
        journal.records_processed = created + updated
        journal.records_created = created
        journal.records_updated = updated
        journal.success = True
        journal.finished_at = timezone.now()
        journal.save()
    except Exception as e:
        logger.exception("NVD sync failed")
        journal.success = False
        journal.error_message = str(e)
        journal.finished_at = timezone.now()
        journal.save()
    return journal


def sync_kev(tenant, triggered_by="manual"):
    journal = SyncJournal.objects.create(tenant=tenant, source="kev", triggered_by=triggered_by)
    try:
        url = "https://www.cisa.gov/sites/default/files/feeds/known_exploited_vulnerabilities.json"
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        data = resp.json()
        created = updated = 0
        for item in data.get("vulnerabilities", []):
            cve_id = item.get("cveID", "")
            fields = {
                "cve_id": cve_id,
                "title": item.get("vulnerabilityName") or cve_id,
                "description_en": item.get("shortDescription", ""),
                "is_kev": True,
                "kev_date_added": item.get("dateAdded") or None,
                "kev_due_date": item.get("dueDate") or None,
                "kev_ransomware": item.get("knownRansomwareCampaignUse", ""),
                "sources": ["kev"],
            }
            _, was_created = _upsert_vuln(tenant, **fields)
            if was_created:
                created += 1
            else:
                updated += 1
        journal.records_processed = created + updated
        journal.records_created = created
        journal.records_updated = updated
        journal.success = True
        journal.finished_at = timezone.now()
        journal.save()
    except Exception as e:
        logger.exception("KEV sync failed")
        journal.success = False
        journal.error_message = str(e)
        journal.finished_at = timezone.now()
        journal.save()
    return journal


def import_bdu_file(tenant, uploaded_file, triggered_by="manual"):
    journal = SyncJournal.objects.create(tenant=tenant, source="bdu", triggered_by=triggered_by)
    name = uploaded_file.name.lower()
    created = updated = 0
    try:
        content = uploaded_file.read()
        rows = []
        if name.endswith(".json"):
            data = json.loads(content.decode("utf-8"))
            rows = data if isinstance(data, list) else data.get("vulnerabilities", data.get("items", []))
        elif name.endswith(".xml"):
            root = ET.fromstring(content)
            for el in root.iter():
                if el.tag.lower().endswith("vulnerability") or el.tag.lower().endswith("item"):
                    rows.append({c.tag.split("}")[-1]: (c.text or "") for c in el})
            if not rows:
                # flat children
                for child in list(root):
                    rows.append({c.tag.split("}")[-1]: (c.text or "") for c in child})
        elif name.endswith(".xlsx") or name.endswith(".xls"):
            wb = load_workbook(filename=io_bytes(content), read_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: row[i] for i in range(len(headers)) if headers[i]})
        else:
            raise ValueError("Unsupported format; use XML / JSON / Excel")

        for row in rows:
            bdu_id = str(row.get("bdu_id") or row.get("id") or row.get("identifier") or "")
            cve_id = str(row.get("cve_id") or row.get("cve") or row.get("cveid") or "")
            desc_ru = str(row.get("description_ru") or row.get("description") or row.get("описание") or "")
            score = row.get("cvss") or row.get("cvss_score") or row.get("base_score")
            try:
                score = float(score) if score not in (None, "") else None
            except (TypeError, ValueError):
                score = None
            fields = {
                "bdu_id": bdu_id,
                "cve_id": cve_id,
                "title": str(row.get("title") or row.get("name") or bdu_id or cve_id),
                "description_ru": desc_ru,
                "cvss_v3_score": score,
                "bdu_raw": row if isinstance(row, dict) else {},
                "sources": ["bdu"],
            }
            _, was_created = _upsert_vuln(tenant, **fields)
            if was_created:
                created += 1
            else:
                updated += 1

        journal.records_processed = created + updated
        journal.records_created = created
        journal.records_updated = updated
        journal.success = True
        journal.finished_at = timezone.now()
        journal.save()
    except Exception as e:
        logger.exception("BDU import failed")
        journal.success = False
        journal.error_message = str(e)
        journal.finished_at = timezone.now()
        journal.save()
    return journal


def io_bytes(content):
    import io

    return io.BytesIO(content)
